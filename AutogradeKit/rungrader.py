#!/usr/bin/env python 3.6
# -*- coding: utf-8 -*-
# @Time    : 2/23/18 13:04
# @Author  : Yam
# @Site    : 
# @File    : rungrader.py
# @Software: PyCharm

import autograde as ag
from openpyxl import Workbook

def insert_excel(sheet, map):
    sheet.cell(row=1, column=1).value = 'ID'
    sheet.cell(row=1, column=2).value = 'Nickname'
    sheet.cell(row=1, column=3).value = 'Date'
    sheet.cell(row=1, column=4).value = 'Time'
    sheet.cell(row=1, column=5).value = 'Score'
    row = 2
    for line in map:
        sheet.cell(row=row, column=1).value = line[0][0]
        sheet.cell(row=row, column=2).value = line[0][1]
        sheet.cell(row=row, column=3).value = line[0][2][0]
        sheet.cell(row=row, column=4).value = line[0][2][1]
        sheet.cell(row=row, column=5).value = line[1]
        row = row + 1

print('Please put logfile into logs.')
logname = 'logs/'+input("Enter log logname:(without suffix)")+'.txt'
date = input("Enter date:(mmdd)")
time = input("Enter time:(hhmm)")
filename = input("Enter excel filename:")

grader = ag.AutoGrader(logname, date, time)
grader.file_read()
# grader.late_punish()
grader.find_best()
# scoremap
# grader.find_late()
# latescoremap
grader.best_check()
# print("------------")
# grader.late_check()

# new file
wb = Workbook()
# new sheet and name
ws = wb.active
ws.title = "BestBeforeDeadline"
insert_excel(ws, grader.scoremap)
wsl = wb.create_sheet(title='BestAfterDeadline')
insert_excel(wsl, grader.latescoremap)
wb.save('outputs/'+filename+'_Score.xlsx')

